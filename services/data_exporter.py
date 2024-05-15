import csv
import json
import pandas as pd
import xml.etree.ElementTree as ET
from database_manager import Database
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


class DataExporter:
    @staticmethod
    def export_to_csv():
        with Database.get_connection() as conn:
            cursor = conn.cursor()
            query = '''SELECT Name, StudyGroup, IDNP, Age, Gender, PersonalEmail, EmergencyContact FROM students
                            JOIN student_info ON students.StudentID = student_info.StudentID;'''
            cursor.execute(query)
            with open('data/students.csv', 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                # Запись заголовков столбцов
                writer.writerow([desc[0] for desc in cursor.description])
                writer.writerows(cursor)
            print("Data exported to CSV")

    @staticmethod
    def export_to_excel():
        with Database.get_connection() as conn:
            cursor = conn.cursor()
            query = '''SELECT Name, StudyGroup, IDNP, Age, Gender, PersonalEmail, EmergencyContact FROM students
                            JOIN student_info ON students.StudentID = student_info.StudentID;'''
            cursor.execute(query)
            columns = [column[0] for column in cursor.description]
            data = cursor.fetchall()

            df = pd.DataFrame.from_records(data, columns=columns)

            # Сохранение DataFrame в Excel
            with pd.ExcelWriter('data/students.xlsx', engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Students')

                # Авто-растяжение столбцов и выравнивание контента по центру
                worksheet = writer.sheets['Students']

                center_aligned_text = Alignment(horizontal='center')
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                for column in df:
                    col_idx = df.columns.get_loc(column) + 1
                    col_letter = get_column_letter(col_idx)
                    column_width = max(df[column].astype(
                        str).map(len).max(), len(column)) + 2
                    worksheet.column_dimensions[col_letter].width = column_width

                    # Применение выравнивания по центру и добавление границы для каждой ячейки в столбце
                    # начинаем с 2, потому что 1-я строка - это заголовки
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = center_aligned_text
                        cell.border = thin_border

                # Добавление границ для заголовков
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}1"]
                    cell.border = thin_border

                print("Data exported to Excel")

    @staticmethod
    def export_to_json():
        with Database.get_connection() as conn:
            cursor = conn.cursor()
            query = '''SELECT Name, StudyGroup, IDNP, Age, Gender, PersonalEmail, EmergencyContact FROM students
                            JOIN student_info ON students.StudentID = student_info.StudentID;'''
            cursor.execute(query)
            columns = [column[0] for column in cursor.description]
            data = cursor.fetchall()

            students_list = []
            for row in data:
                student = {col: val for col, val in zip(columns, row)}
                students_list.append(student)

            with open('data/students.json', 'w', encoding='utf-8') as file:
                json.dump(students_list, file, ensure_ascii=False, indent=4)
            print("Data exported to JSON")

    @staticmethod
    def export_to_xml():
        def indent(elem, level=0):
            i = "\n" + level * "  "
            if len(elem):
                if not elem.text or not elem.text.strip():
                    elem.text = i + "  "
                if not elem.tail or not elem.tail.strip():
                    elem.tail = i
                for elem in elem:
                    indent(elem, level + 1)
                if not elem.tail or not elem.tail.strip():
                    elem.tail = i
            else:
                if level and (not elem.tail or not elem.tail.strip()):
                    elem.tail = i

        with Database.get_connection() as conn:
            cursor = conn.cursor()
            query = '''SELECT Name, StudyGroup, IDNP, Age, Gender, PersonalEmail, EmergencyContact FROM students
                            JOIN student_info ON students.StudentID = student_info.StudentID;'''
            cursor.execute(query)
            columns = [column[0] for column in cursor.description]
            data = cursor.fetchall()

            root = ET.Element('Students')
            for row in data:
                student = ET.SubElement(root, 'Student')
                for col, val in zip(columns, row):
                    col_elem = ET.SubElement(student, col)
                    col_elem.text = str(val)

            indent(root)
            tree = ET.ElementTree(root)
            tree.write('data/students.xml', encoding='utf-8',
                       xml_declaration=True)
            print("Data exported to XML")
